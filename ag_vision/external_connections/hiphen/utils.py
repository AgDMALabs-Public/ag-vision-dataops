import json
import os
import re
import sys
import time
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from urllib.parse import urlparse, parse_qs
from datetime import datetime as dt
from uuid import uuid4
from open_aglabs.image.models import AgImageModel
import glob

from ag_vision.constants import paths as pth
import pandas as pd

import requests

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl is required: pip install openpyxl")
    sys.exit(1)

DEFAULT_BASE_URL = "https://api.hiphen-cloverfield.com"
ORTHO_FILENAME = "orthomosaic.tif"


class CloverfieldError(Exception):
    """Base exception for Cloverfield API errors."""


class AuthenticationError(CloverfieldError):
    """Raised when authentication with the Cloverfield API fails."""


# ═══════════════════════════════════════════════════════════════════════════
# TOKEN MANAGER — automatic refresh
# ═══════════════════════════════════════════════════════════════════════════

class TokenManager:
    """
    Manages the Cloverfield access token with automatic refresh.

    Tokens expire after 300 seconds. A refresh is triggered proactively
    once the token is older than 240 seconds, providing a 60-second safety
    margin before expiration.
    """

    def __init__(self, base_url: str, client_id: str, client_secret: str):
        self.base_url = base_url
        self.client_id = client_id
        self.client_secret = client_secret
        self._token = None
        self._obtained_at = 0
        self._expires_in = 300

    def _fetch_token(self):
        url = f"{self.base_url}/token"
        resp = requests.post(url, timeout=30, json={
            "clientId": self.client_id,
            "clientSecret": self.client_secret,
        })
        if resp.status_code != 200:
            print(f"Authentication error ({resp.status_code}): {resp.text}")
            raise AuthenticationError(f"Authentication error ({resp.status_code}): {resp.text}")

        data = resp.json()
        self._token = data.get("accessToken")
        self._expires_in = data.get("expiresIn", 300)
        self._obtained_at = time.time()

        if not self._token:
            print("Missing 'accessToken' field in response:", data)
            raise AuthenticationError("Authentication error (Missing token)")

    def get_headers(self) -> dict:
        """Return request headers with a valid bearer token, refreshing it if necessary."""
        age = time.time() - self._obtained_at
        # Refresh if no token yet, or if the token is within 60 seconds of expiration
        if self._token is None or age > (self._expires_in - 60):
            if self._token is not None:
                print("Refreshing access token ...")
            else:
                print("Authenticating ...")
            self._fetch_token()
            print(f"Access token obtained (expires in {self._expires_in}s)")

        return {
            "Authorization": f"Bearer {self._token}",
            "Accept": "application/json",
        }


# ═══════════════════════════════════════════════════════════════════════════
# Class Hypehn Data
# ═══════════════════════════════════════════════════════════════════════════ -
class HiphenData:
    """
    Manages the Cloverfield access token with automatic refresh.

    Tokens expire after 300 seconds. A refresh is triggered proactively
    once the token is older than 240 seconds, providing a 60-second safety
    margin before expiration.
    """

    def __init__(self, token_mgr: TokenManager, year: int = None, country: str = None, crop: str = None,
                 season: str = None,
                 flight_date: str = None, site_name: str = None, field_name: str = None,
                 location_name: str = None, project_dir: str = None, trial_name: str = None,
                 mission_name: str = None, camera: str = None):
        self.token_mgr = token_mgr
        self.contracts = None
        self.contract_id = None
        self.sites = None
        self.site_id = None
        self.site_id = None
        self.mission_dates = None
        self.flight_date = flight_date
        self.ortho_file = None
        self.plot_files = None
        self.raw_results = None
        self.tabular_results = None
        self.geojson_results = None
        self.plot_boundaries = None
        self.project_dir = project_dir
        self.trial_name = trial_name
        self.year = year
        self.country = country
        self.crop = crop
        self.season = season
        self.crop_season = None
        self.site_name = site_name
        self.field_name = field_name
        self.location_name = location_name
        self.mission_name = mission_name
        self.camera = camera
        self.mission_dir = None
        self.data_summary = None

    def set_crop_season(self):
        self.crop_season = pth.season_code(year=self.year,
                                           country=self.country,
                                           crop=self.crop,
                                           time_of_year=self.season)

    def set_mission_dir(self):
        assert self.project_dir is not None, f"set the project dir"
        assert self.site_name is not None, f"set the site name"
        assert self.trial_name is not None, f"set the trial name"
        assert self.crop_season is not None, f"set the crop season"
        assert self.field_name is not None, f"set the field name"
        assert self.location_name is not None, f"set the location name"
        assert self.mission_name is not None, f"set the missions name"

        loc_path = pth.location_path(project=self.project_dir,
                                     site=self.site_name,
                                     trial=self.trial_name,
                                     season=self.crop_season,
                                     field=self.field_name,
                                     location=self.location_name)

        self.mission_dir = pth.drone_mission_dir(location_path=loc_path,
                                                 mission_name=self.mission_name)

    def set_flight_date(self, flight_date):
        self.flight_date = flight_date
        # self.flight_date = pd.to_datetime(flight_date)

    def list_contracts(self):
        self.contracts = list_contracts(token_mgr=self.token_mgr)

    def set_contract_id(self, index):
        self.contract_id = self.contracts['data'][index]['id']

    def list_sites(self):
        self.sites = list_contract_sites(token_mgr=self.token_mgr,
                                         contract_id=self.contract_id)

    def set_site_id(self, index):
        self.site_id = self.sites['data'][index]['id']

    def get_site_info(self):
        self.site_info = get_site_info(site_id=self.site_id,
                                       token_mgr=self.token_mgr)

        if self.site_info:
            self.mission_dates = self.site_info['missions']

    def get_plot_image_list(self):
        assert self.flight_date is not None, f"The flight date need to tbe set"
        assert self.site_id is not None, f"The site ID needs to be set"

        self.plot_files = fetch_plot_images_for_date(site_id=self.site_id,
                                                     date=self.flight_date,
                                                     token_mgr=self.token_mgr)

    def get_ortho_img_path(self):
        assert self.flight_date is not None, f"The flight date needs to be set"
        assert self.site_id is not None, f"The site ID needs to be set."

        self.ortho_file = fetch_ortho_urls_for_date(site_id=self.site_id,
                                                    date=self.flight_date,
                                                    token_mgr=self.token_mgr)

    def get_hiphen_plot_metrics(self):
        self.raw_results = fetch_all_results(site_id=self.site_id,
                                             token_mgr=self.token_mgr)

    def generate_results_dataframe(self):
        assert self.raw_results is not None, f"Raw Results need to be pulled"
        assert self.flight_date is not None, f"Need to set the flight_date"

        self.tabular_results = results_to_dataframe(results=self.raw_results,
                                                    flight_date=self.flight_date)

    def generate_results_geojson(self):
        assert self.raw_results is not None, f"Raw Results need to be pulled"
        assert self.flight_date is not None, f"Need to set the flight_date"

        self.geojson_results = create_geojson(results=self.raw_results,
                                              flight_date=self.flight_date)

    def save_tabular_results(self):
        assert self.mission_name is not None, f"Need to set the mission_dir"
        assert self.flight_date is not None, f"Need to set the flight date"
        assert self.camera is not None, f"Need to set the camera"

        dst_path = pth.drone_results_path(mission_dir=self.mission_dir,
                                          flight_date=self.flight_date,
                                          camera=self.camera,
                                          file_name='hiphen_results.csv')

        dir_name = os.path.dirname(dst_path)
        os.makedirs(dir_name, exist_ok=True)

        self.tabular_results.to_csv(dst_path)

    def save_geojson_results(self):
        assert self.mission_name is not None, f"Need to set the mission_dir"
        assert self.flight_date is not None, f"Need to set the flight date"
        assert self.camera is not None, f"Need to set the camera"

        dst_path = pth.drone_results_path(mission_dir=self.mission_dir,
                                          flight_date=self.flight_date,
                                          camera=self.camera,
                                          file_name='hiphen_results.geojson')

        dir_name = os.path.dirname(dst_path)
        os.makedirs(dir_name, exist_ok=True)

        write_geojson(data=self.geojson_results,
                      filepath=dst_path)

    def download_plot_images(self):
        assert self.flight_date is not None, f"The flight date is not set"
        assert self.camera is not None, f"The camera is not set"

        datetime = str(dt.now())
        for plot in self.plot_files:
            for img in plot['images']:
                filename = img['fileName']
                ext = os.path.splitext(filename)[1]
                img_id = str(uuid4())
                dst_path = pth.drone_flight_plot_image_path(mission_dir=self.mission_dir,
                                                            flight_date=self.flight_date,
                                                            datetime=datetime,
                                                            camera=self.camera,
                                                            plot_id=str(plot['id']),
                                                            image_name=img_id + str(ext))

                dir_name = os.path.dirname(dst_path)
                os.makedirs(dir_name, exist_ok=True)

                _download_single_image(img_url=img['url'],
                                       dest_path=dst_path)

                metadata = {
                    "path": dst_path,
                    "id": img_id,
                    "device": "drone",
                    "type": "original",
                    "protocol_properties": {
                    },
                    "trial_properties": {
                        "name": self.trial_name
                    },
                    "camera_properties": {
                    },
                    "location_properties": {
                        "id": plot['id'],
                        "name": plot['id'],
                        "admin_level_0": self.country,
                        "site": self.site_name,
                        "field": self.field_name,
                        "location": self.location_name
                    },
                    "acquisition_properties": {
                        "date": self.flight_date,
                        "light_source": "natural",
                        "setting": "field"
                    },
                    "image_quality": {
                    },
                    "agronomic_properties": {
                        "crop_type": self.crop,
                    }
                }
                aim = AgImageModel(**metadata)

                validated_json = aim.model_dump_json()
                metadata_path = dst_path.replace(ext, '.json')
                # Write the string directly to your file path
                with open(metadata_path, "w", encoding="utf-8") as f:
                    f.write(validated_json)

    def download_ortho_images(self):
        assert self.flight_date is not None, "Flight date needs to be set"
        assert self.camera is not None, "The camera needs to be set"

        assert len(self.ortho_file) == 1, f'The len of the orth files is {len(self.ortho_file)} and it should be 1'

        ortho = self.ortho_file[0]

        dst_path = pth.drone_flight_orthomosaic_path(mission_dir=self.mission_dir,
                                                     flight_date=self.flight_date,
                                                     method='hiphen',
                                                     ortho_date=str(dt.now()),
                                                     camera=self.camera,
                                                     image_name=ortho['fileName'])

        dir_name = os.path.dirname(dst_path)
        os.makedirs(dir_name, exist_ok=True)

        download_orthoimage(images=self.ortho_file,
                            dest_path=Path(dst_path))

    def generate_contract_site_mission_table(self):
        df_list = []
        for c_idx, contract in enumerate(self.contracts['data']):
            self.set_contract_id(index=c_idx)
            self.list_sites()
            for s_idx, site in enumerate(self.sites['data']):
                self.set_site_id(index=s_idx)
                self.get_site_info()
                print(self.site_info)
                df = pd.DataFrame({'flight_date': self.site_info['missions']})
                df['country'] = self.site_info["country"]
                df['crop'] = self.site_info["crop"]
                df['location'] = self.site_info['displayName']
                df['contract_id'] = str(contract['id'])
                df['contract_idx'] = c_idx
                df['contract'] = str(contract['name'])
                df['site_idx'] = s_idx
                df['site_id'] = str(site['id'])
                df['site'] = str(site['name'])
                df_list.append(df)

        out_df = pd.concat(df_list)
        out_df.reset_index(drop=True)

        self.data_summary = out_df

    def add_plot_dir_to_summary_table(self, volume):
        assert self.data_summary is not None
        self.data_summary = self.data_summary.reset_index(drop=True)
        for idx, row in self.data_summary.iterrows():
            self.data_summary.loc[idx, 'plot_dir'] = os.path.join(volume,
                                                                  row['contract'],
                                                                  '*',
                                                                  '*',
                                                                  row['location'],
                                                                  row['location'],
                                                                  'drone',
                                                                  '*',
                                                                  row['flight_date'],
                                                                  'plot_image',
                                                                  '*',
                                                                  'rgb',
                                                                  '*',
                                                                  '*.webp')

            self.data_summary['plot_dir'] =  self.data_summary['plot_dir'].apply(lambda x: x.lower())

    def count_plot_images(self):
        assert self.data_summary is not None
        assert 'plot_dir' in self.data_summary.columns
        for idx, row in self.data_summary.iterrows():
            count = glob.glob(row['plot_dir'])
            self.data_summary.loc[idx, 'plot_img_count'] = len(count)


# ═══════════════════════════════════════════════════════════════════════════


# HELPERS
# ═══════════════════════════════════════════════════════════════════════════

def sanitize_name(name: str) -> str:
    """
    Sanitize a string to be safe for filenames by replacing invalid characters.

    :param name: The original string to sanitize.
    :return: A sanitized string safe for use in filenames.
    :rtype: str
    """
    name = re.sub(r'[<>:"/\\|?*]', '_', name)
    name = re.sub(r'\s+', '_', name)
    return name.strip('_.')


def extract_cursor(next_link: str) -> str | None:
    """
    Extract the 'page[next]' cursor value from a pagination link.

    :param next_link: The URL from the 'next' pagination link.
    :return: The cursor value for the next page, or None if not available.
    :rtype: str or None
    """
    if not next_link:
        return None
    if next_link.startswith("http"):
        qs = parse_qs(urlparse(next_link).query)
        return qs.get("page[next]", [None])[0]
    return next_link


def ensure_string_decimal(value) -> str:
    """
    Convert a float to a string, or return an empty string if None.

    :param value: The value to convert.
    :return: A string representation of the float, or an empty
                string if the value is None.
    :rtype: str
    """
    if value is None:
        return ""
    return str(value)


# ═══════════════════════════════════════════════════════════════════════════
# API — Site info
# ═══════════════════════════════════════════════════════════════════════════
def list_contracts(token_mgr: TokenManager) -> dict:
    """
    Retrieve site information, including display name, crop type, plot count, and missions.

    :param site_id: The UUID of the site to retrieve.
    :param token_mgr: An instance of TokenManager for authenticated requests.
    :return: A dictionary containing site information.
    :rtype: dict
    """
    url = f"{token_mgr.base_url}/contracts"
    print("Retrieving the contract lists ...")
    resp = requests.get(url,
                        headers=token_mgr.get_headers(),
                        timeout=30)
    if resp.status_code != 200:
        print(f"Site retrieval error ({resp.status_code}): {resp.text}")
        raise CloverfieldError(f"Site retrieval error ({resp.status_code}): {resp.text}")
    contracts = resp.json()
    # display = site.get("displayName") or site.get("name") or site_id
    # missions = site.get("missions", [])
    print(f" - Contracts: {contracts}")
    return contracts


def list_contract_sites(token_mgr, contract_id):
    """
    Fetches a specific contract by ID.
    The returned data inherently includes all sites and upload templates for the contract.
    """
    # The endpoint requires the specific contract ID in the URL path
    url = f"{token_mgr.base_url}/contracts/{contract_id}/sites"
    print(f"Calling {url}")

    # Use the access token generated previously as a Bearer token
    resp = requests.get(url,
                        headers=token_mgr.get_headers(),
                        timeout=30)
    if resp.status_code != 200:
        print(f"Site retrieval error ({resp.status_code}): {resp.text}")
        raise CloverfieldError(f"Site retrieval error ({resp.status_code}): {resp.text}")

    # Parse the JSON response
    sites = resp.json()
    print(f"The sites are {sites}")

    return sites


def get_site_info(site_id: str, token_mgr: TokenManager) -> dict:
    """
    Retrieve site information, including display name, crop type, plot count, and missions.

    :param site_id: The UUID of the site to retrieve.
    :param token_mgr: An instance of TokenManager for authenticated requests.
    :return: A dictionary containing site information.
    :rtype: dict
    """
    url = f"{token_mgr.base_url}/sites/{site_id}"
    print("Retrieving site information ...")
    resp = requests.get(url, headers=token_mgr.get_headers(), timeout=30)
    if resp.status_code != 200:
        print(f"Site retrieval error ({resp.status_code}): {resp.text}")
        raise CloverfieldError(f"Site retrieval error ({resp.status_code}): {resp.text}")
    site = resp.json()
    display = site.get("displayName") or site.get("name") or site_id
    missions = site.get("missions", [])
    print(f" - Site: {display}")
    print(f" - Missions ({len(missions)}): {', '.join(missions)}")
    return site


# ═══════════════════════════════════════════════════════════════════════════
# API — Paginated results
# ═══════════════════════════════════════════════════════════════════════════

def fetch_all_results(site_id: str, token_mgr: TokenManager) -> list:
    """
    Retrieve all results for a site, handling pagination automatically.

    :param site_id: The UUID of the site to retrieve results for.
    :param token_mgr: An instance of TokenManager for authenticated requests.
    :return: A list of all result records for the site.
    :rtype: list
    """
    url = f"{token_mgr.base_url}/sites/{site_id}/results"
    all_results = []
    page = 1
    params = {}

    print("Fetching results ...")
    while True:
        print(f" - Page {page} ...", end=" ")
        resp = requests.get(url, headers=token_mgr.get_headers(), params=params, timeout=30)
        if resp.status_code != 200:
            print(f"\nError ({resp.status_code}): {resp.text}")
            raise CloverfieldError(f"Results retrieval error ({resp.status_code}): {resp.text}")

        body = resp.json()
        data = body.get("data", [])
        meta = body.get("meta", {})
        all_results.extend(data)
        print(f"{len(data)} result(s) (total: {len(all_results)})")

        cursor = extract_cursor(meta.get("links", {}).get("next"))
        if cursor:
            params = {"page[next]": cursor}
            page += 1
        else:
            break

    print(f" - {len(all_results)} result(s) retrieved")
    return all_results


# ═══════════════════════════════════════════════════════════════════════════
# API — Orthoimages (for a specific date)
# ═══════════════════════════════════════════════════════════════════════════

def fetch_ortho_urls_for_date(site_id: str, date: str, token_mgr: TokenManager) -> list:
    """
    Retrieve fresh presigned URLs for orthoimages on a specific date.
    Returns a list of image descriptors in the form [{url, fileName}, ...].

    :param site_id: The UUID of the site to retrieve orthoimages for.
    :param date: The date (YYYY-MM-DD) to filter orthoimages by.
    :param token_mgr: An instance of TokenManager for authenticated requests.
    :return: A list of image descriptors with 'url' and 'fileName' keys
    :rtype: list
    """
    url = f"{token_mgr.base_url}/sites/{site_id}/orthoimages"
    resp = requests.get(url,
                        headers=token_mgr.get_headers(),
                        params={"date": date},
                        timeout=30)
    if resp.status_code != 200:
        return []
    body = resp.json()
    for item in body.get("data", []):
        if item.get("date") == date:
            return item.get("images", [])
    return []


def download_orthoimage(images: list, dest_path: Path) -> bool:
    """
    Download the orthoimage from the provided list of image descriptors.
    It prioritizes images with "plain" in the filename or those ending with ".tif
    for better quality, but will fall back to the first available image if necessary.

    :param images: A list of image descriptors with 'url' and 'fileName' keys.
    :param dest_path: The local file path to save the downloaded orthoimage to.
    :return: True if the download was successful, False otherwise.
    :rtype: bool
    """
    target = None
    for img in images:
        fn = img.get("fileName", "")
        if "plain" in fn.lower() or fn.endswith(".tif"):
            target = img
            break

    if not target and images:
        target = images[0]
    if not target or not target.get("url"):
        print("No orthoimage available")
        return False

    print("Downloading orthoimage ...", end=" ", flush=True)
    try:
        resp = requests.get(target["url"], stream=True, timeout=600)
        if resp.status_code != 200:
            print(f"({resp.status_code})")
            return False
        downloaded = 0
        with open(dest_path, "wb") as f:
            for chunk in resp.iter_content(chunk_size=65536):
                f.write(chunk)
                downloaded += len(chunk)
        size_mb = downloaded / (1024 * 1024)
        print(f"{size_mb:.1f} MB")
        return True
    except Exception as e:
        print(f"Error on orthoimage download {e}")
        return False


# ═══════════════════════════════════════════════════════════════════════════
# API — Plot images for a specific date
# ═══════════════════════════════════════════════════════════════════════════
def _download_single_image(img_url, dest_path):
    resp = requests.get(img_url, stream=True, timeout=300)
    if resp.status_code == 200:
        with open(dest_path, "wb") as f:
            for chunk in resp.iter_content(chunk_size=8192):
                f.write(chunk)
        return True
    return False


def fetch_plot_images_for_date(site_id: str, date: str, token_mgr: TokenManager) -> list:
    """
    Retrieve all plot image descriptors for a specific date, handling pagination.
    Each item in the returned list corresponds to a plot and contains its associated images.

    :param base_url: The base URL of the Cloverfield API.
    :param site_id: The UUID of the site to retrieve plot images for.
    :param date: The date (YYYY-MM-DD) to filter plot images by.
    :param token_mgr: An instance of TokenManager for authenticated requests.
    :return: A list of plot items, each containing an 'images' key with a
            list of image descriptors ({url, fileName}).
    :rtype: list
    """
    url = f"{token_mgr.base_url}/sites/{site_id}/images"
    print(f"Fetching data from {url}")
    all_items = []
    params = {"date": date}

    while True:
        resp = requests.get(url, headers=token_mgr.get_headers(), params=params, timeout=30)
        if resp.status_code != 200:
            print(f"Plot images retrieval error ({resp.status_code})")
            print(resp)
            return all_items
        body = resp.json()
        data = body.get("data", [])
        meta = body.get("meta", {})
        all_items.extend(data)

        cursor = extract_cursor(meta.get("links", {}).get("next"))
        if cursor:
            params = {"date": date, "page[next]": cursor}
        else:
            break
    return all_items


def download_plot_images(site_id: str, date: str,
                         token_mgr: TokenManager, plots_dir: Path) -> int:
    """
    Download all plot images for a specific date, saving them to the provided directory.
    It retrieves the plot image index first, then downloads each image in parallel.

    :param site_id: The UUID of the site to retrieve plot images for.
    :param date: The date (YYYY-MM-DD) to filter plot images by.
    :param token_mgr: An instance of TokenManager for authenticated requests.
    :param plots_dir: The local directory to save the downloaded plot images to.
    :return: The total number of images successfully downloaded.
    :rtype: int
    """
    print("Fetching plot image index ...", end=" ", flush=True)
    items = fetch_plot_images_for_date(site_id, date, token_mgr)
    if not items:
        print("no plot images found")
        return 0

    total_images = sum(len(item.get("images", [])) for item in items)
    print(f"{total_images} plot image(s) across {len(items)} plot(s)")
    plots_dir.mkdir(parents=True, exist_ok=True)

    downloaded = 0
    errors = 0
    with ThreadPoolExecutor(max_workers=8) as pool:
        futures = []
        for item in items:
            plot_id = item.get("id", "unknown")
            for img in item.get("images", []):
                img_url = img.get("url", "")
                if not img_url:
                    continue
                file_name = img.get("fileName", "") or f"plot_{plot_id}.webp"
                dest_path = plots_dir / file_name
                # dest_path = pth.drone_flight_plot_image_path(mission_dir=)
                futures.append(pool.submit(_download_single_image, img_url, dest_path))

        for future in as_completed(futures):
            if not future.exception():
                downloaded += 1
            else:
                errors += 1

            if downloaded > 0 and downloaded % 100 == 0:
                print(f"{downloaded}/{total_images} downloaded", flush=True)

    print(f"{downloaded} image(s) downloaded" +
          (f", {errors} error(s)" if errors else ""))
    return downloaded


# ═══════════════════════════════════════════════════════════════════════════
# RESULT FLATTENING
# ═══════════════════════════════════════════════════════════════════════════

def build_flat_rows(results: list) -> tuple:
    """
    Build a flat list of rows for Excel export, along with the complete
    set of columns and geometries. Each row corresponds to a single result
    and includes all top-level properties and traits as columns.

    :param results: A list of result records from the Cloverfield API.
    :return: A tuple containing:
                - A list of dictionaries representing rows for Excel export.
                - A list of all column names (properties + traits).
                - A dictionary mapping result IDs to their geometries (if available).
    :rtype: tuple
    """
    rows = []
    client_cols = []
    client_seen = set()
    trait_cols = []
    trait_seen = set()
    geometries = {}

    for r in results:
        props_data = r.get("properties", {}).get("data", {})
        traits = r.get("properties", {}).get("traits", [])
        plot_id = r.get("id", "")
        row = {}

        row["_result_id"] = plot_id

        for key, value in props_data.items():
            row[key] = ensure_string_decimal(value) if isinstance(value, float) else value
            if key not in client_seen:
                client_cols.append(key)
                client_seen.add(key)

        for trait in traits:
            api_name = trait.get("name", "unknown")
            row[api_name] = ensure_string_decimal(trait.get("value"))
            if api_name not in trait_seen:
                trait_cols.append(api_name)
                trait_seen.add(api_name)

        rows.append(row)
        geom = r.get("geometry")
        if geom and plot_id:
            geometries[plot_id] = geom

    all_columns = client_cols + trait_cols
    return rows, all_columns, geometries


# ═══════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ═══════════════════════════════════════════════════════════════════════════
def results_to_dataframe(results: list, flight_date) -> pd.DataFrame:
    """
    Convert the results to a pandas DataFrame.

    :param rows: A list of dictionaries representing rows for the DataFrame.
    :param columns: A list of column names to include in the DataFrame.
    :return: A pandas DataFrame with the results.
    """
    # Create DataFrame from the list of dictionaries
    # Only include columns that are in the columns list
    grouped = defaultdict(list)
    for r in results:
        grouped[r.get("date", "unknown")].append(r)

    df_list = []
    for date_key, records in sorted(grouped.items()):
        if date_key not in [flight_date]:
            continue
        # a) Excel
        rows, columns, geometries = build_flat_rows(records)
        df = pd.DataFrame(rows, columns=columns)
        df['date'] = date_key
        df_list.append(df)

    return pd.concat(df_list)


def write_excel(rows: list, columns: list, filepath: Path):
    """
    Write the results to an Excel file.

    :param rows: A list of dictionaries representing rows for Excel export.
    :param columns: A list of column names to include in the Excel file.
    :param filepath: The local file path to save the Excel file to.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    header_font = Font(name="Calibri", bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="top")

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.alignment = header_align

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name))

    if rows and columns:
        last_col = get_column_letter(len(columns))
        ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

    wb.save(filepath)


# ═══════════════════════════════════════════════════════════════════════════
# GEOJSON EXPORT
# ═══════════════════════════════════════════════════════════════════════════
def create_geojson(results: list, flight_date: str):
    """
    Write the results to a GeoJSON file.

    :param rows: A list of dictionaries representing rows for GeoJSON export.
    :param columns: A list of column names to include in the GeoJSON file.
    :param geometries: A dictionary mapping result IDs to their geometries.
    """
    grouped = defaultdict(list)
    for r in results:
        grouped[r.get("date", "unknown")].append(r)

    features = []
    for date_key, records in sorted(grouped.items()):
        if date_key not in [flight_date]:
            continue
        rows, columns, geometries = build_flat_rows(records)

        for row in rows:
            _result_id = row.get("_result_id", "")
            geometry = geometries.get(_result_id)
            plot_id = row.get("plot_id", row.get("client_id", _result_id))
            properties = {}
            for col in columns:
                val = row.get(col)
                if col in ("Y", "X") and val is not None:
                    try:
                        properties[col] = int(val)
                    except (ValueError, TypeError):
                        properties[col] = val
                elif val is not None:
                    properties[col] = str(val)
                else:
                    properties[col] = val

            features.append({
                "type": "Feature",
                "id": str(plot_id),
                "properties": properties,
                "geometry": geometry
            })

    return {"type": "FeatureCollection", "features": features}


def write_geojson(data: dict, filepath: str):
    """

    :param filepath: The local file path to save the GeoJSON file to.
    """

    filepath = Path(filepath)
    dir = os.path.dirname(filepath)
    os.makedirs(dir, exist_ok=True)

    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


# ═══════════════════════════════════════════════════════════════════════════
# MAIN PIPELINE
# ═══════════════════════════════════════════════════════════════════════════

def process_and_export(results: list,
                       base_url: str, site_id: str,
                       token_mgr: TokenManager,
                       site_name: str,
                       mission_dir: str,
                       camera: str,
                       flight_date: str,
                       download_ortho: bool = False,
                       download_images: bool = False):
    """
    Process and export results to Excel and GeoJSON formats.
    Results are grouped by their 'date' property, and each date's results are exported
    to a separate folder named {site_name}_{date}/. Optionally, orthomosaic images
    and per-plot images can also be downloaded for each date if requested.

    :param results: A list of result records from the Cloverfield API.
    :param base_url: The base URL of the Cloverfield API.
    :param site_id: The UUID of the site to retrieve additional data for.
    :param token_mgr: An instance of TokenManager for authenticated requests.
    :param site_name: The display name of the site, used for folder naming.
    :param mission_dir: The local directory to save the exported files to.
    :param download_ortho: Whether to download orthomosaic images for each date.
    :param download_images: Whether to download per-plot images for each date.
    """
    if not results:
        print("No results to export.")
        return

    grouped = defaultdict(list)
    for r in results:
        grouped[r.get("date", "unknown")].append(r)

    print(f"\n{len(grouped)} mission date(s): {', '.join(sorted(grouped.keys()))}")
    safe_name = sanitize_name(site_name)

    for date_key, records in sorted(grouped.items()):
        # a) Excel
        rows, columns, geometries = build_flat_rows(records)
        # xlsx_path = folder_path / f"{basename}.xlsx"
        xlsx_path = pth.drone_results_path(mission_dir=mission_dir,
                                           flight_date=flight_date,
                                           camera=camera,
                                           file_name='hiphen_results.xlsx')

        write_excel(rows, columns, xlsx_path)
        print(f"{xlsx_path} ({len(records)} plots)")

        # b) GeoJSON
        geojson_path = pth.drone_results_path(mission_dir=mission_dir,
                                              flight_date=flight_date,
                                              camera=camera,
                                              file_name='hiphen_results.geojson')

        write_geojson(rows, columns, geometries, geojson_path)
        print(f"{geojson_path}")

        # c) Orthoimage — fresh presigned URLs are fetched immediately before download
        if download_ortho:
            ortho_images = fetch_ortho_urls_for_date(
                base_url, site_id, date_key, token_mgr)
            if ortho_images:
                ortho_dest = pth.drone_flight_orthomosaic_path(mission_dir=mission_dir,
                                                               flight_date=flight_date,
                                                               method='hiphen',
                                                               ortho_date=dt.now(),
                                                               camera=camera,
                                                               image_name='orthomosaic.tif')
                download_orthoimage(ortho_images, ortho_dest)
            else:
                print(f"No orthoimage available for {date_key}")

        # d) Plot images — token is refreshed automatically by TokenManager
        if download_images:
            # plots_dir = folder_path / "Plots"
            plots_dir = pth.drone_flight_plot_image_dir(mission_dir=mission_dir,
                                                        flight_date=flight_date,
                                                        datetime=dt.now(),
                                                        camera=camera,
                                                        plot_id='', )
            download_plot_images(site_id, date_key, token_mgr, plots_dir)

    print(f"\nResults ready in: {str(output_dir)}")
