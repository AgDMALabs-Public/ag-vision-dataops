#!/usr/bin/env python3
"""
# This was the example code given by the Cloverfield API team.

#
Cloverfield Results Downloader
==============================

Downloads results by date from the Cloverfield API.
This script is compatible with Python 3.10+ and requires the `requests` and `openpyxl` libraries.

Output structure per date:
    {site_name}_{date}/
    ├── {site_name}_{date}.xlsx
    ├── {site_name}_{date}.geojson
    ├── orthomosaic.tif (only if explicitly requested whith --download-ortho)
    └── Plots/
        └── plot_{plot_id}_*.webp (only if explicitly requested with --download-images)

Usage:
    python download_cloverfield_results.py \
        YOUR_SITE_UUID \
        YOUR_OUTPUT_DIRECTORY \
        --client-id YOUR_CLIENT_ID \
        --client-secret YOUR_CLIENT_SECRET
"""

import argparse
import json
import os
import re
import sys
import time
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from urllib.parse import urlparse, parse_qs

import requests

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl is required: pip install openpyxl")
    sys.exit(1)


DEFAULT_BASE_URL = "https://api.hiphen-cloverfield.com"
ORTHO_FILENAME = "orthomosaic.tif"


class CloverfieldError(Exception):
    """Base exception for Cloverfield API errors."""


class AuthenticationError(CloverfieldError):
    """Raised when authentication with the Cloverfield API fails."""


# ═══════════════════════════════════════════════════════════════════════════
# TOKEN MANAGER — automatic refresh
# ═══════════════════════════════════════════════════════════════════════════

class TokenManager:
    """
    Manages the Cloverfield access token with automatic refresh.

    Tokens expire after 300 seconds. A refresh is triggered proactively
    once the token is older than 240 seconds, providing a 60-second safety
    margin before expiration.
    """
    def __init__(self, base_url: str, client_id: str, client_secret: str):
        self.base_url = base_url
        self.client_id = client_id
        self.client_secret = client_secret
        self._token = None
        self._obtained_at = 0
        self._expires_in = 300

    def _fetch_token(self):
        url = f"{self.base_url}/token"
        resp = requests.post(url, timeout=30, json={
            "clientId": self.client_id,
            "clientSecret": self.client_secret,
        })
        if resp.status_code != 200:
            print(f"Authentication error ({resp.status_code}): {resp.text}")
            raise AuthenticationError(f"Authentication error ({resp.status_code}): {resp.text}")

        data = resp.json()
        self._token = data.get("accessToken")
        self._expires_in = data.get("expiresIn", 300)
        self._obtained_at = time.time()

        if not self._token:
            print("Missing 'accessToken' field in response:", data)
            raise AuthenticationError("Authentication error (Missing token)")

    def get_headers(self) -> dict:
        """Return request headers with a valid bearer token, refreshing it if necessary."""
        age = time.time() - self._obtained_at
        # Refresh if no token yet, or if the token is within 60 seconds of expiration
        if self._token is None or age > (self._expires_in - 60):
            if self._token is not None:
                print("Refreshing access token ...")
            else:
                print("Authenticating ...")
            self._fetch_token()
            print(f"Access token obtained (expires in {self._expires_in}s)")

        return {
            "Authorization": f"Bearer {self._token}",
            "Accept": "application/json",
        }


# ═══════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════

def sanitize_name(name: str) -> str:
    """
    Sanitize a string to be safe for filenames by replacing invalid characters.

    :param name: The original string to sanitize.
    :return: A sanitized string safe for use in filenames.
    :rtype: str
    """
    name = re.sub(r'[<>:"/\\|?*]', '_', name)
    name = re.sub(r'\s+', '_', name)
    return name.strip('_.')


def extract_cursor(next_link: str) -> str | None:
    """
    Extract the 'page[next]' cursor value from a pagination link.

    :param next_link: The URL from the 'next' pagination link.
    :return: The cursor value for the next page, or None if not available.
    :rtype: str or None
    """
    if not next_link:
        return None
    if next_link.startswith("http"):
        qs = parse_qs(urlparse(next_link).query)
        return qs.get("page[next]", [None])[0]
    return next_link


def ensure_string_decimal(value) -> str:
    """
    Convert a float to a string, or return an empty string if None.

    :param value: The value to convert.
    :return: A string representation of the float, or an empty
                string if the value is None.
    :rtype: str
    """
    if value is None:
        return ""
    return str(value)


# ═══════════════════════════════════════════════════════════════════════════
# API — Site info
# ═══════════════════════════════════════════════════════════════════════════

def get_site_info(base_url: str, site_id: str, token_mgr: TokenManager) -> dict:
    """
    Retrieve site information, including display name, crop type, plot count, and missions.

    :param base_url: The base URL of the Cloverfield API.
    :param site_id: The UUID of the site to retrieve.
    :param token_mgr: An instance of TokenManager for authenticated requests.
    :return: A dictionary containing site information.
    :rtype: dict
    """
    url = f"{base_url}/sites/{site_id}"
    print("Retrieving site information ...")
    resp = requests.get(url, headers=token_mgr.get_headers(), timeout=30)
    if resp.status_code != 200:
        print(f"Site retrieval error ({resp.status_code}): {resp.text}")
        raise CloverfieldError(f"Site retrieval error ({resp.status_code}): {resp.text}")
    site = resp.json()
    display = site.get("displayName") or site.get("name") or site_id
    missions = site.get("missions", [])
    print(f" - Site: {display}")
    print(f" - Missions ({len(missions)}): {', '.join(missions)}")
    return site


# ═══════════════════════════════════════════════════════════════════════════
# API — Paginated results
# ═══════════════════════════════════════════════════════════════════════════

def fetch_all_results(base_url: str, site_id: str, token_mgr: TokenManager) -> list:
    """
    Retrieve all results for a site, handling pagination automatically.

    :param base_url: The base URL of the Cloverfield API.
    :param site_id: The UUID of the site to retrieve results for.
    :param token_mgr: An instance of TokenManager for authenticated requests.
    :return: A list of all result records for the site.
    :rtype: list
    """
    url = f"{base_url}/sites/{site_id}/results"
    all_results = []
    page = 1
    params = {}

    print("Fetching results ...")
    while True:
        print(f" - Page {page} ...", end=" ")
        resp = requests.get(url, headers=token_mgr.get_headers(), params=params, timeout=30)
        if resp.status_code != 200:
            print(f"\nError ({resp.status_code}): {resp.text}")
            raise CloverfieldError(f"Results retrieval error ({resp.status_code}): {resp.text}")

        body = resp.json()
        data = body.get("data", [])
        meta = body.get("meta", {})
        all_results.extend(data)
        print(f"{len(data)} result(s) (total: {len(all_results)})")

        cursor = extract_cursor(meta.get("links", {}).get("next"))
        if cursor:
            params = {"page[next]": cursor}
            page += 1
        else:
            break

    print(f" - {len(all_results)} result(s) retrieved")
    return all_results


# ═══════════════════════════════════════════════════════════════════════════
# API — Orthoimages (for a specific date)
# ═══════════════════════════════════════════════════════════════════════════

def fetch_ortho_urls_for_date(base_url: str, site_id: str, date: str,
                              token_mgr: TokenManager) -> list:
    """
    Retrieve fresh presigned URLs for orthoimages on a specific date.
    Returns a list of image descriptors in the form [{url, fileName}, ...].

    :param base_url: The base URL of the Cloverfield API.
    :param site_id: The UUID of the site to retrieve orthoimages for.
    :param date: The date (YYYY-MM-DD) to filter orthoimages by.
    :param token_mgr: An instance of TokenManager for authenticated requests.
    :return: A list of image descriptors with 'url' and 'fileName' keys
    :rtype: list
    """
    url = f"{base_url}/sites/{site_id}/orthoimages"
    resp = requests.get(url, headers=token_mgr.get_headers(), params={"date": date}, timeout=30)
    if resp.status_code != 200:
        return []
    body = resp.json()
    for item in body.get("data", []):
        if item.get("date") == date:
            return item.get("images", [])
    return []


def download_orthoimage(images: list, dest_path: Path) -> bool:
    """
    Download the orthoimage from the provided list of image descriptors.
    It prioritizes images with "plain" in the filename or those ending with ".tif
    for better quality, but will fall back to the first available image if necessary.

    :param images: A list of image descriptors with 'url' and 'fileName' keys.
    :param dest_path: The local file path to save the downloaded orthoimage to.
    :return: True if the download was successful, False otherwise.
    :rtype: bool
    """
    target = None
    for img in images:
        fn = img.get("fileName", "")
        if "plain" in fn.lower() or fn.endswith(".tif"):
            target = img
            break
    if not target and images:
        target = images[0]
    if not target or not target.get("url"):
        print("No orthoimage available")
        return False

    print("Downloading orthoimage ...", end=" ", flush=True)
    try:
        resp = requests.get(target["url"], stream=True, timeout=600)
        if resp.status_code != 200:
            print(f"({resp.status_code})")
            return False
        downloaded = 0
        with open(dest_path, "wb") as f:
            for chunk in resp.iter_content(chunk_size=65536):
                f.write(chunk)
                downloaded += len(chunk)
        size_mb = downloaded / (1024 * 1024)
        print(f"{size_mb:.1f} MB")
        return True
    except Exception as e:
        print(f"Error on orthoimage download {e}")
        return False


# ═══════════════════════════════════════════════════════════════════════════
# API — Plot images for a specific date
# ═══════════════════════════════════════════════════════════════════════════
def _download_single_image(img_url, dest_path):
    resp = requests.get(img_url, stream=True, timeout=300)
    if resp.status_code == 200:
        with open(dest_path, "wb") as f:
            for chunk in resp.iter_content(chunk_size=8192):
                f.write(chunk)
        return True
    return False


def fetch_plot_images_for_date(base_url: str, site_id: str, date: str,
                               token_mgr: TokenManager) -> list:
    """
    Retrieve all plot image descriptors for a specific date, handling pagination.
    Each item in the returned list corresponds to a plot and contains its associated images.

    :param base_url: The base URL of the Cloverfield API.
    :param site_id: The UUID of the site to retrieve plot images for.
    :param date: The date (YYYY-MM-DD) to filter plot images by.
    :param token_mgr: An instance of TokenManager for authenticated requests.
    :return: A list of plot items, each containing an 'images' key with a
            list of image descriptors ({url, fileName}).
    :rtype: list
    """
    url = f"{base_url}/sites/{site_id}/images"
    all_items = []
    params = {"date": date}

    while True:
        resp = requests.get(url, headers=token_mgr.get_headers(), params=params, timeout=30)
        if resp.status_code != 200:
            print(f"Plot images retrieval error ({resp.status_code})")
            return all_items
        body = resp.json()
        data = body.get("data", [])
        meta = body.get("meta", {})
        all_items.extend(data)

        cursor = extract_cursor(meta.get("links", {}).get("next"))
        if cursor:
            params = {"date": date, "page[next]": cursor}
        else:
            break
    return all_items


def download_plot_images(base_url: str, site_id: str, date: str,
                        token_mgr: TokenManager, plots_dir: Path) -> int:
    """
    Download all plot images for a specific date, saving them to the provided directory.
    It retrieves the plot image index first, then downloads each image in parallel.

    :param base_url: The base URL of the Cloverfield API.
    :param site_id: The UUID of the site to retrieve plot images for.
    :param date: The date (YYYY-MM-DD) to filter plot images by.
    :param token_mgr: An instance of TokenManager for authenticated requests.
    :param plots_dir: The local directory to save the downloaded plot images to.
    :return: The total number of images successfully downloaded.
    :rtype: int
    """
    print("Fetching plot image index ...", end=" ", flush=True)
    items = fetch_plot_images_for_date(base_url, site_id, date, token_mgr)
    if not items:
        print("no plot images found")
        return 0

    total_images = sum(len(item.get("images", [])) for item in items)
    print(f"{total_images} plot image(s) across {len(items)} plot(s)")
    plots_dir.mkdir(parents=True, exist_ok=True)

    downloaded = 0
    errors = 0
    with ThreadPoolExecutor(max_workers=8) as pool:
        futures = []
        for item in items:
            plot_id = item.get("id", "unknown")
            for img in item.get("images", []):
                img_url = img.get("url", "")
                if not img_url:
                    continue
                file_name = img.get("fileName", "") or f"plot_{plot_id}.webp"
                dest_path = plots_dir / file_name
                futures.append(pool.submit(_download_single_image, img_url, dest_path))

        for future in as_completed(futures):
            if not future.exception():
                downloaded += 1
            else:
                errors += 1

            if downloaded > 0 and downloaded % 100 == 0:
                print(f"{downloaded}/{total_images} downloaded", flush=True)

    print(f"{downloaded} image(s) downloaded" +
          (f", {errors} error(s)" if errors else ""))
    return downloaded


# ═══════════════════════════════════════════════════════════════════════════
# RESULT FLATTENING
# ═══════════════════════════════════════════════════════════════════════════

def build_flat_rows(results: list) -> tuple:
    """
    Build a flat list of rows for Excel export, along with the complete
    set of columns and geometries. Each row corresponds to a single result
    and includes all top-level properties and traits as columns.

    :param results: A list of result records from the Cloverfield API.
    :return: A tuple containing:
                - A list of dictionaries representing rows for Excel export.
                - A list of all column names (properties + traits).
                - A dictionary mapping result IDs to their geometries (if available).
    :rtype: tuple
    """
    rows = []
    client_cols = []
    client_seen = set()
    trait_cols = []
    trait_seen = set()
    geometries = {}

    for r in results:
        props_data = r.get("properties", {}).get("data", {})
        traits = r.get("properties", {}).get("traits", [])
        plot_id = r.get("id", "")
        row = {}

        row["_result_id"] = plot_id

        for key, value in props_data.items():
            row[key] = ensure_string_decimal(value) if isinstance(value, float) else value
            if key not in client_seen:
                client_cols.append(key)
                client_seen.add(key)

        for trait in traits:
            api_name = trait.get("name", "unknown")
            row[api_name] = ensure_string_decimal(trait.get("value"))
            if api_name not in trait_seen:
                trait_cols.append(api_name)
                trait_seen.add(api_name)

        rows.append(row)
        geom = r.get("geometry")
        if geom and plot_id:
            geometries[plot_id] = geom

    all_columns = client_cols + trait_cols
    return rows, all_columns, geometries


# ═══════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ═══════════════════════════════════════════════════════════════════════════

def write_excel(rows: list, columns: list, filepath: Path):
    """
    Write the results to an Excel file.

    :param rows: A list of dictionaries representing rows for Excel export.
    :param columns: A list of column names to include in the Excel file.
    :param filepath: The local file path to save the Excel file to.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    header_font = Font(name="Calibri", bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="top")

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.alignment = header_align

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name))

    if rows and columns:
        last_col = get_column_letter(len(columns))
        ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

    wb.save(filepath)


# ═══════════════════════════════════════════════════════════════════════════
# GEOJSON EXPORT
# ═══════════════════════════════════════════════════════════════════════════

def write_geojson(rows: list, columns: list, geometries: dict, filepath: Path):
    """
    Write the results to a GeoJSON file.

    :param rows: A list of dictionaries representing rows for GeoJSON export.
    :param columns: A list of column names to include in the GeoJSON file.
    :param geometries: A dictionary mapping result IDs to their geometries.
    :param filepath: The local file path to save the GeoJSON file to.
    """
    features = []
    for row in rows:
        _result_id = row.get("_result_id", "")
        geometry = geometries.get(_result_id)
        plot_id = row.get("plot_id", row.get("client_id", _result_id))
        properties = {}
        for col in columns:
            val = row.get(col)
            if col in ("Y", "X") and val is not None:
                try:
                    properties[col] = int(val)
                except (ValueError, TypeError):
                    properties[col] = val
            elif val is not None:
                properties[col] = str(val)
            else:
                properties[col] = val

        features.append({
            "type": "Feature",
            "id": str(plot_id),
            "properties": properties,
            "geometry": geometry
        })

    geojson = {"type": "FeatureCollection", "features": features}
    filepath = Path(filepath)
    filepath.parent.mkdir(parents=True, exist_ok=True)

    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(geojson, f, indent=2, ensure_ascii=False)


# ═══════════════════════════════════════════════════════════════════════════
# MAIN PIPELINE
# ═══════════════════════════════════════════════════════════════════════════

def process_and_export(results: list,
                       base_url: str, site_id: str, token_mgr: TokenManager,
                       site_name: str, output_dir: str,
                       download_ortho: bool = False, download_images: bool = False):
    """
    Process and export results to Excel and GeoJSON formats.
    Results are grouped by their 'date' property, and each date's results are exported
    to a separate folder named {site_name}_{date}/. Optionally, orthomosaic images
    and per-plot images can also be downloaded for each date if requested.

    :param results: A list of result records from the Cloverfield API.
    :param base_url: The base URL of the Cloverfield API.
    :param site_id: The UUID of the site to retrieve additional data for.
    :param token_mgr: An instance of TokenManager for authenticated requests.
    :param site_name: The display name of the site, used for folder naming.
    :param output_dir: The local directory to save the exported files to.
    :param download_ortho: Whether to download orthomosaic images for each date.
    :param download_images: Whether to download per-plot images for each date.
    """
    if not results:
        print("No results to export.")
        return

    grouped = defaultdict(list)
    for r in results:
        grouped[r.get("date", "unknown")].append(r)

    print(f"\n{len(grouped)} mission date(s): {', '.join(sorted(grouped.keys()))}")
    safe_name = sanitize_name(site_name)

    for date_key, records in sorted(grouped.items()):
        basename = f"{safe_name}_{date_key}"
        folder_path = Path(output_dir) / basename
        folder_path.mkdir(parents=True, exist_ok=True)

        print(f"\n{basename}/")

        # a) Excel
        rows, columns, geometries = build_flat_rows(records)
        xlsx_path = folder_path / f"{basename}.xlsx"
        write_excel(rows, columns, xlsx_path)
        print(f"{basename}.xlsx ({len(records)} plots)")

        # b) GeoJSON
        geojson_path = folder_path / f"{basename}.geojson"
        write_geojson(rows, columns, geometries, geojson_path)
        print(f"geojson/{basename}.geojson")

        # c) Orthoimage — fresh presigned URLs are fetched immediately before download
        if download_ortho:
            ortho_images = fetch_ortho_urls_for_date(
                base_url, site_id, date_key, token_mgr)
            if ortho_images:
                ortho_dest = folder_path / ORTHO_FILENAME
                download_orthoimage(ortho_images, ortho_dest)
            else:
                print(f"No orthoimage available for {date_key}")

        # d) Plot images — token is refreshed automatically by TokenManager
        if download_images:
            plots_dir = folder_path / "Plots"
            download_plot_images(base_url, site_id, date_key, token_mgr, plots_dir)

    print(f"\nResults ready in: {str(output_dir)}")


# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main():
    """
    Parse command-line arguments and execute the main processing pipeline.
    """
    parser = argparse.ArgumentParser(
        description="Download results from the Cloverfield API, organized by mission date."
    )
    parser.add_argument(
        "site_id",
        help="Cloverfield site UUID (v4)"
        )
    parser.add_argument(
        "output_dir",
        help="Local output directory"
    )
    parser.add_argument(
        "--client-id",
        required=True,
        help="Cloverfield Client ID"
    )
    parser.add_argument(
        "--client-secret",
        required=True,
        help="Cloverfield Client Secret"
    )
    parser.add_argument(
        "--base-url",
        default=DEFAULT_BASE_URL,
        help=f"API base URL (default: {DEFAULT_BASE_URL})"
        )
    parser.add_argument(
        "--download-ortho",
        action="store_true",
        help="Download orthomosaic images"
        )
    parser.add_argument(
        "--download-images",
        action="store_true",
        help="Download per-plot images"
        )

    args = parser.parse_args()

    try:
        os.makedirs(args.output_dir, exist_ok=True)

        # TokenManager handles automatic token refresh throughout the session
        token_mgr = TokenManager(args.base_url, args.client_id, args.client_secret)

        site_info = get_site_info(args.base_url, args.site_id, token_mgr)
        site_name = site_info.get("name") or site_info.get("displayName") or f"site_{args.site_id}"

        results = fetch_all_results(args.base_url, args.site_id, token_mgr)

        process_and_export(
            results,
            args.base_url, args.site_id, token_mgr, site_name, args.output_dir,
            download_ortho=args.download_ortho, download_images=args.download_images,
        )
    except CloverfieldError as e:
        print(f"Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
